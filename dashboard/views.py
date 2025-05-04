from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count
from django.utils import timezone
import csv
from datetime import datetime, timedelta
from forms.models import (
    T3PersonelVeriler, 
    GonulluDurumVeriler, 
    GonulluSorunVeriler, 
    SorumluVeriler,
    SistemAyarlari,
    T3PersonelAtama
)
from forms.views import role_required
from accounts.views import log_user_action
from accounts.models import User
from django.shortcuts import get_object_or_404
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.db.models import Sum, F
from django.core.paginator import Paginator
from django.template.defaulttags import register
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
import re

# Template filtreleri
@register.filter
def get_item(dictionary, key):
    """Dictionary'den değer almak için template filtresi"""
    return dictionary.get(key, 0)

@register.filter
def dictsort(dictionary, key):
    """Dictionary'yi belirli bir key'e göre sıralamak için template filtresi"""
    return dict(sorted(dictionary.items()))

@register.filter
def dictitem(dictionary, key):
    """Dictionary'den belirli bir key'e göre değer almak için template filtresi"""
    if key in dictionary:
        return dictionary[key]
    return {}

@login_required
@role_required(['izleyici', 'admin'])
def dashboard_home(request):
    """Dashboard ana sayfası"""
    log_user_action(request, 'Dashboard Ana Sayfası Görüntülendi', 'Dashboard Home')

    # Son 7 günlük verileri getir
    son_7_gun = timezone.now().date() - timedelta(days=7)

    t3_veriler_sayisi = T3PersonelVeriler.objects.filter(submitteddate__gte=son_7_gun).count()


    gonullu_durum_sayisi = GonulluDurumVeriler.objects.filter(submitteddate__gte=son_7_gun).count()
    gonullu_sorun_sayisi = GonulluSorunVeriler.objects.filter(submitteddate__gte=son_7_gun).count()
    sorumlu_veriler_sayisi = SorumluVeriler.objects.filter(submitteddate__gte=son_7_gun).count()


    # Toplam sipariş sayıları
    toplam_personel_siparis = SorumluVeriler.objects.filter(submitteddate__gte=son_7_gun).aggregate(Sum('personel_yemek_siparis'))['personel_yemek_siparis__sum'] or 0
    toplam_taseron_siparis = SorumluVeriler.objects.filter(submitteddate__gte=son_7_gun).aggregate(Sum('taseron_yemek_siparis'))['taseron_yemek_siparis__sum'] or 0
    

    toplam_t3_siparis = (
        T3PersonelVeriler.objects
        .filter(submitteddate__gte=son_7_gun)
        .aggregate(toplam=Sum(F('ogle_yemegi') + F('aksam_yemegi') + F('lunchbox')))['toplam'] or 0
    )

    # Koordinatörlük tablosu için veri hazırlama
    # Seçilen gün (varsayılan olarak 1. gün)
    secilen_gun = request.GET.get('gun', '1. Gün')
    
    # Tüm koordinatörlük listesi
    koordinatorlukler = [
        'Selçuk Bey VIP', 'VIP Salon', 'Yönetim Ofisi / Kriz Masası', 'Vakıf Standı Yönetici Odası',
        'Dinamik Alan', 'TSK', 'Gönüllü - Bursiyer', 'Basın', 'TRT Kulis Ana Sahne / Muhabir',
        'Arter / To Do / Kurumsal Ofis / Heysemist', 'Pilot Event', 'T3 Ofis', 
        'Teknofest Robolig Yarışması', 'Teknofest KKTC Araştırma Yarışması', 'Bilim Pavyonu'
    ]
    
    # Tüm günler
    gunler = ['1. Gün', '2. Gün', '3. Gün', '4. Gün']
    
    # Koordinatörlük verilerini hesapla
    koordinatorluk_veri_sayilari = {}
    
    # Her koordinatörlük için seçilen günde kaç veri girildiğini hesapla
    for koordinatorluk in koordinatorlukler:
        # Durum verilerini say
        durum_sayisi = GonulluDurumVeriler.objects.filter(
            alan=koordinatorluk, 
            gun=secilen_gun
        ).count()
        
        # Sorun verilerini say
        sorun_sayisi = GonulluSorunVeriler.objects.filter(
            alan=koordinatorluk, 
            gun=secilen_gun
        ).count()
        
        # Toplam veri sayısı
        toplam_veri = durum_sayisi + sorun_sayisi
        
        koordinatorluk_veri_sayilari[koordinatorluk] = toplam_veri

    # Sistem ayarlarını al
    try:
        veri_guncelleme_son_saat = int(SistemAyarlari.objects.get(anahtar='veri_guncelleme_son_saat').deger)
    except (SistemAyarlari.DoesNotExist, ValueError):
        veri_guncelleme_son_saat = 14  # Varsayılan değer

    try:
        veri_guncelleme_son_dakika = int(SistemAyarlari.objects.get(anahtar='veri_guncelleme_son_dakika').deger)
    except (SistemAyarlari.DoesNotExist, ValueError):
        veri_guncelleme_son_dakika = 0  # Varsayılan değer

    context = {
        't3_veriler_sayisi': t3_veriler_sayisi,
        'gonullu_durum_sayisi': gonullu_durum_sayisi,
        'gonullu_sorun_sayisi': gonullu_sorun_sayisi,
        'sorumlu_veriler_sayisi': sorumlu_veriler_sayisi,
        'toplam_personel_siparis': toplam_personel_siparis,
        'toplam_taseron_siparis': toplam_taseron_siparis,
        'toplam_t3_siparis': toplam_t3_siparis,
        'veri_guncelleme_son_saat': veri_guncelleme_son_saat,
        'veri_guncelleme_son_dakika': veri_guncelleme_son_dakika,
        'secilen_gun': secilen_gun,
        'gunler': gunler,
        'koordinatorlukler': koordinatorlukler,
        'koordinatorluk_veri_sayilari': koordinatorluk_veri_sayilari,
    }

    return render(request, 'dashboard/home.html', context)

@login_required
@role_required(['izleyici', 'admin'])
def t3personel_dashboard(request):
    """T3 personel verileri dashboard"""
    log_user_action(request, 'T3 Personel Dashboard Görüntülendi', 'T3 Personel Dashboard')

    # Filtreleme parametreleri
    baslangic_tarihi = request.GET.get('baslangic_tarihi')
    bitis_tarihi = request.GET.get('bitis_tarihi')
    koordinatorluk = request.GET.get('koordinatorluk')
    birim = request.GET.get('birim')

    # Temel sorgu
    veriler = T3PersonelVeriler.objects.all().order_by('-submitteddate', '-submittedtime')

    # Filtreleri uygula
    if baslangic_tarihi:
        veriler = veriler.filter(submitteddate__gte=baslangic_tarihi)
    if bitis_tarihi:
        veriler = veriler.filter(submitteddate__lte=bitis_tarihi)
    if koordinatorluk:
        veriler = veriler.filter(koordinatorluk__icontains=koordinatorluk)
    if birim:
        veriler = veriler.filter(birim__icontains=birim)

    # CSV indirme
    # Excel indirme

    if 'csv' in request.GET:



        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "T3 Personel Verileri"



        # Başlıklar

        headers = ['TC', 'İsim', 'Soyisim', 'Koordinatörlük', 'Birim',

                'Öğle Yemeği', 'Akşam Yemeği', 'Lunchbox', 'Toplam', 'Tarih', 'Saat']

        ws.append(headers)



        # Veriler

        for veri in veriler:

            toplam = veri.ogle_yemegi + veri.aksam_yemegi + veri.lunchbox

            ws.append([

                veri.kisi.tc,

                veri.kisi.isim,

                veri.kisi.soyisim,

                veri.koordinatorluk,

                veri.birim,

                veri.ogle_yemegi,

                veri.aksam_yemegi,

                veri.lunchbox,

                toplam,

                veri.submitteddate.strftime('%Y-%m-%d'),

                str(veri.submittedtime)

            ])



        # Excel dosyasını bellekte oluştur

        output = BytesIO()

        wb.save(output)

        output.seek(0)



        # HTTP yanıtı olarak döndür

        response = HttpResponse(

            output,

            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        )

        response['Content-Disposition'] = 'attachment; filename="t3personel_veriler.xlsx"'

        return response

    # Koordinatörlük ve birim listelerini al
    koordinatorlukler = T3PersonelVeriler.objects.values_list('koordinatorluk', flat=True).distinct()
    birimler = T3PersonelVeriler.objects.values_list('birim', flat=True).distinct()

    context = {
        'veriler': veriler,
        'koordinatorlukler': koordinatorlukler,
        'birimler': birimler,
        'filtreler': {
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
            'koordinatorluk': koordinatorluk,
            'birim': birim,
        }
    }

    return render(request, 'dashboard/t3personel.html', context)

@login_required
@role_required(['izleyici', 'admin'])
def gonullu_durum_dashboard(request):
    """Gönüllü durum verileri dashboard"""
    log_user_action(request, 'Gönüllü Durum Dashboard Görüntülendi', 'Gönüllü Durum Dashboard')

    # Filtreleme parametreleri
    baslangic_tarihi = request.GET.get('baslangic_tarihi')
    bitis_tarihi = request.GET.get('bitis_tarihi')
    alan = request.GET.get('alan')
    gun = request.GET.get('gun')
    catering_durum = request.GET.get('catering_durum')
    kisi_isim = request.GET.get('kisi_isim')
    aciklama = request.GET.get('aciklama')
    
    # Temel sorgu
    veriler = GonulluDurumVeriler.objects.all().order_by('-submitteddate', '-submittedtime')

    # Filtreleri uygula
    if baslangic_tarihi:
        veriler = veriler.filter(submitteddate__gte=baslangic_tarihi)
    if bitis_tarihi:
        veriler = veriler.filter(submitteddate__lte=bitis_tarihi)
    if alan:
        veriler = veriler.filter(alan__icontains=alan)
    if gun:
        veriler = veriler.filter(gun__icontains=gun)
    if catering_durum:
        veriler = veriler.filter(catering_durum=catering_durum)
    if kisi_isim:
        veriler = veriler.filter(kisi__isim__icontains=kisi_isim) | veriler.filter(kisi__soyisim__icontains=kisi_isim)
    if aciklama:
        veriler = veriler.filter(aciklama__icontains=aciklama)

    # CSV indirme
    if 'csv' in request.GET:
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="gonullu_durum_veriler.csv"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['İsim', 'Gün', 'Saat', 'Alan', 'Catering Durumu', 'Catering Ürünleri', 'Açıklama', 'Fotoğraf', 'Tarih', 'Saat'])
        for veri in veriler:
            writer.writerow([
                veri.kisi.get_full_name(),
                veri.gun,
                veri.saat,
                veri.alan,
                veri.catering_durum,
                ', '.join(veri.catering_urunleri) if veri.catering_urunleri else '',
                veri.aciklama if veri.aciklama else '-',
                'Var' if veri.fotograf else 'Yok',
                veri.submitteddate,
                veri.submittedtime,
            ])

        return response

    # Pagination
    paginator = Paginator(veriler, 50)  # Her sayfada 50 kayıt
    page = request.GET.get('page')
    veriler_paginated = paginator.get_page(page)
    
    # Her veri için fotografları prefetch et
    for veri in veriler_paginated:
        veri.fotograflar_list = veri.fotograflar.all()

    # Alan listesini al
    alanlar = GonulluDurumVeriler.objects.values_list('alan', flat=True).distinct()
    
    # Gün seçeneklerini al
    gunler = GonulluDurumVeriler.objects.values_list('gun', flat=True).distinct()
    
    # Catering durumu seçenekleri
    catering_durumlari = [('var', 'Catering Var'), ('yok', 'Catering Yok')]

    context = {
        'veriler': veriler_paginated,
        'alanlar': alanlar,
        'gunler': gunler,
        'catering_durumlari': catering_durumlari,
        'filtreler': {
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
            'alan': alan,
            'gun': gun,
            'catering_durum': catering_durum,
            'kisi_isim': kisi_isim,
            'aciklama': aciklama,
        }
    }

    return render(request, 'dashboard/gonullu_durum.html', context)

@login_required
@role_required(['izleyici', 'admin'])
def gonullu_sorun_dashboard(request):
    """Gönüllü sorun verileri dashboard"""
    log_user_action(request, 'Gönüllü Sorun Dashboard Görüntülendi', 'Gönüllü Sorun Dashboard')

    # Filtreleme parametreleri
    baslangic_tarihi = request.GET.get('baslangic_tarihi')
    bitis_tarihi = request.GET.get('bitis_tarihi')
    alan = request.GET.get('alan')
    sorun_tipi = request.GET.get('sorun_tipi')
    sorun_seviyesi = request.GET.get('sorun_seviyesi')
    gun = request.GET.get('gun')
    kisi_isim = request.GET.get('kisi_isim')

    # Temel sorgu
    veriler = GonulluSorunVeriler.objects.all().order_by('-submitteddate', '-submittedtime')

    # Filtreleri uygula
    if baslangic_tarihi:
        veriler = veriler.filter(submitteddate__gte=baslangic_tarihi)
    if bitis_tarihi:
        veriler = veriler.filter(submitteddate__lte=bitis_tarihi)
    if alan:
        veriler = veriler.filter(alan__icontains=alan)
    if sorun_tipi:
        veriler = veriler.filter(sorun_tipi=sorun_tipi)
    if sorun_seviyesi:
        veriler = veriler.filter(sorun_seviyesi=sorun_seviyesi)
    if gun:
        veriler = veriler.filter(gun__icontains=gun)
    if kisi_isim:
        veriler = veriler.filter(kisi__isim__icontains=kisi_isim) | veriler.filter(kisi__soyisim__icontains=kisi_isim)

    # CSV indirme
    if 'csv' in request.GET:
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="gonullu_sorun_veriler.csv"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['TC', 'İsim', 'Soyisim', 'Gün', 'Saat', 'Alan', 'Sorun Tipi', 'Sorun Seviyesi', 'Açıklama', 'Tarih', 'Saat'])

        for veri in veriler:
            writer.writerow([
                veri.kisi.tc,
                veri.kisi.isim,
                veri.kisi.soyisim,
                veri.gun,
                veri.saat,
                veri.alan,
                veri.sorun_tipi,
                veri.sorun_seviyesi,
                veri.aciklama,
                veri.submitteddate.strftime('%Y-%m-%d'),
                veri.submittedtime.strftime('%H:%M:%S')
            ])

        return response
        
    # Pagination
    paginator = Paginator(veriler, 50)  # Her sayfada 50 kayıt
    page = request.GET.get('page')
    veriler_paginated = paginator.get_page(page)
    
    # Her veri için fotografları prefetch et
    for veri in veriler_paginated:
        veri.fotograflar_list = veri.fotograflar.all()

    # Alan listesini al
    alanlar = GonulluSorunVeriler.objects.values_list('alan', flat=True).distinct()
    
    # Gün listesini al
    gunler = GonulluSorunVeriler.objects.values_list('gun', flat=True).distinct()
    
    # Sorun tipleri ve seviyeleri için choices'ları al
    sorun_tipleri = [choice[0] for choice in GonulluSorunVeriler.SORUN_TIPI_CHOICES]
    sorun_seviyeleri = [choice[0] for choice in GonulluSorunVeriler.SORUN_SEVIYESI_CHOICES]

    context = {
        'veriler': veriler_paginated,
        'alanlar': alanlar,
        'gunler': gunler,
        'sorun_tipleri': sorun_tipleri,
        'sorun_seviyeleri': sorun_seviyeleri,
        'filtreler': {
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
            'alan': alan,
            'sorun_tipi': sorun_tipi,
            'sorun_seviyesi': sorun_seviyesi,
            'gun': gun,
            'kisi_isim': kisi_isim
        }
    }

    return render(request, 'dashboard/gonullu_sorun.html', context)

@login_required
@role_required(['izleyici', 'admin', 'sorumlu'])
def sorumlu_dashboard(request):
    """Sorumlu verileri dashboard"""
    log_user_action(request, 'Sorumlu Dashboard Görüntülendi', 'Sorumlu Dashboard')

    # Filtreleme parametreleri
    baslangic_tarihi = request.GET.get('baslangic_tarihi')
    bitis_tarihi = request.GET.get('bitis_tarihi')

    # Temel sorgu
    veriler = SorumluVeriler.objects.all().order_by('-submitteddate', '-submittedtime')

    # Filtreleri uygula
    if baslangic_tarihi:
        veriler = veriler.filter(submitteddate__gte=baslangic_tarihi)
    if bitis_tarihi:
        veriler = veriler.filter(submitteddate__lte=bitis_tarihi)

    # CSV indirme
    if 'csv' in request.GET:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sorumlu_veriler.csv"'

        writer = csv.writer(response)
        writer.writerow(['TC', 'İsim', 'Soyisim', 'Gün', 'Personel Yemek Siparişi', 'Taşeron Yemek Siparişi', 'Tarih', 'Saat'])

        for veri in veriler:
            writer.writerow([
                veri.kisi.tc,
                veri.kisi.isim,
                veri.kisi.soyisim,
                veri.gun,
                veri.personel_yemek_siparis,
                veri.taseron_yemek_siparis,
                veri.submitteddate,
                veri.submittedtime
            ])

        return response

    # Toplam sipariş sayıları
    toplam_personel_siparis = veriler.aggregate(Sum('personel_yemek_siparis'))['personel_yemek_siparis__sum'] or 0
    toplam_taseron_siparis = veriler.aggregate(Sum('taseron_yemek_siparis'))['taseron_yemek_siparis__sum'] or 0

    context = {
        'veriler': veriler,
        'toplam_personel_siparis': toplam_personel_siparis,
        'toplam_taseron_siparis': toplam_taseron_siparis,
    }

    return render(request, 'dashboard/sorumlu.html', context)

@login_required
def sistem_ayarlari_guncelle(request):
    """Sistem ayarlarını güncelleme view'ı"""
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Bu işlemi yapmaya yetkiniz yok.')
        return redirect('dashboard:home')

    if request.method == 'POST':
        veri_guncelleme_son_saat = request.POST.get('veri_guncelleme_son_saat', '14')
        veri_guncelleme_son_dakika = request.POST.get('veri_guncelleme_son_dakika', '0')

        # Değerleri kontrol et
        try:
            saat = int(veri_guncelleme_son_saat)
            dakika = int(veri_guncelleme_son_dakika)

            if not (0 <= saat <= 23 and 0 <= dakika <= 59):
                raise ValueError("Geçersiz saat veya dakika değeri")

            # Saat ayarını kaydet
            SistemAyarlari.objects.update_or_create(
                anahtar='veri_guncelleme_son_saat',
                defaults={'deger': str(saat), 'aciklama': 'T3 personel verilerinin güncellenebileceği son saat'}
            )

            # Dakika ayarını kaydet
            SistemAyarlari.objects.update_or_create(
                anahtar='veri_guncelleme_son_dakika',
                defaults={'deger': str(dakika), 'aciklama': 'T3 personel verilerinin güncellenebileceği son dakika'}
            )

            messages.success(request, 'Sistem ayarları başarıyla güncellendi.')
        except ValueError:
            messages.error(request, 'Geçersiz saat veya dakika değeri girdiniz.')

    return redirect('dashboard:home')

@login_required
@role_required(['admin'])
def t3personel_atama_ekle(request):
    """T3 personel ataması ekleme sayfası"""
    log_user_action(request, 'T3 Personel Atama Ekleme Sayfası Görüntülendi', 'T3 Personel Atama')

    # T3 personel rolüne sahip kullanıcıları getir
    users = User.objects.filter(role='t3personel')

    # Sabit koordinatörlük listesi
    koordinatorlukler = [
        "Bilişim Koordinatörlüğü",
        "Bursiyer Koordinatörlüğü",
        "Deneyap Koordinatörlüğü",
        "Eğitim Ar-Ge",
        "Fuar Koordinatörlüğü",
        "Girişim Koordinatörlüğü",
        "İdari İşler Koordinatörlüğü",
        "Kurumsal İletişim Koordinatörlüğü",
        "Kurumsal Yapılanma Koordinatörlüğü",
        "Mimari Tasarım Koordinatörlüğü",
        "Satış ve Pazarlama Koordinatörlüğü",
        "Operasyon Koordinatörlüğü",
        "Ulaşım Koordinatörlüğü",
        "Yarışmalar Koordinatörlüğü"
    ]

    # Veritabanındaki diğer koordinatörlükleri de ekle
    db_koordinatorlukler = T3PersonelAtama.objects.values_list('koordinatorluk', flat=True).distinct()
    for k in db_koordinatorlukler:
        if k not in koordinatorlukler:
            koordinatorlukler.append(k)

    # Mevcut atamaları getir
    atamalar = T3PersonelAtama.objects.all().order_by('kisi__isim', 'kisi__soyisim', 'koordinatorluk', 'birim')

    if request.method == 'POST':
        # Form verilerini al
        kisi_ids = request.POST.getlist('kisi')
        koordinatorlukler = request.POST.getlist('koordinatorluk')
        birimler = request.POST.getlist('birim')
        coffee_breaks = request.POST.getlist('coffee_break')
        
        # Her bir satır için işlem yap
        basarili_kayit = 0
        hata_kayit = 0

        for i in range(len(kisi_ids)):
            if i < len(koordinatorlukler) and i < len(birimler):
                try:
                    kisi = get_object_or_404(User, id=kisi_ids[i])
                    koordinatorluk = koordinatorlukler[i]
                    birim = birimler[i]
                    coffee_break = True if f'{i}' in coffee_breaks else False

                    # Boş değer kontrolü
                    if not koordinatorluk or not birim:
                        hata_kayit += 1
                        continue

                    # Aynı kayıt var mı kontrolü
                    if T3PersonelAtama.objects.filter(kisi=kisi, koordinatorluk=koordinatorluk, birim=birim).exists():
                        hata_kayit += 1
                        continue

                    # Yeni atama oluştur
                    T3PersonelAtama.objects.create(
                        kisi=kisi,
                        koordinatorluk=koordinatorluk,
                        birim=birim,
                        coffee_break=coffee_break
                    )
                    basarili_kayit += 1
                except Exception as e:
                    hata_kayit += 1
                    print(f"Hata: {str(e)}")

        if basarili_kayit > 0:
            messages.success(request, f'{basarili_kayit} adet T3 personel ataması başarıyla eklendi.')

        if hata_kayit > 0:
            messages.warning(request, f'{hata_kayit} adet atama eklenirken hata oluştu.')

        return redirect('dashboard:t3personel_atama_ekle')

    context = {
        'users': users,
        'koordinatorlukler': koordinatorlukler,
        'atamalar': atamalar,
    }

    return render(request, 'dashboard/t3personel_atama_ekle.html', context)

@login_required
@role_required(['admin'])
def get_koordinatorluk_for_user(request, user_id):
    """Kullanıcının mevcut koordinatörlüğünü getir"""
    try:
        # Kullanıcının en son atamasını bul
        atama = T3PersonelAtama.objects.filter(kisi_id=user_id).order_by('-id').first()

        if atama:
            return JsonResponse({
                'success': True,
                'koordinatorluk': atama.koordinatorluk
            })
        else:
            return JsonResponse({
                'success': True,
                'koordinatorluk': None
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@role_required(['izleyici', 'admin'])
def gonullu_durum_raporu(request):
    """Gönüllü durum raporu sayfası"""
    log_user_action(request, 'Gönüllü Durum Raporu Görüntülendi', 'Gönüllü Durum Raporu')
    
    # Seçilen gün (varsayılan olarak 1. Gün)
    secilen_gun = request.GET.get('gun', '1. Gün')
    
    # Koordinatörlük/alan listesi - sütunlar olacak
    alanlar = [
        'Arter / To Do / Kurumsal Ofis / Heysemist',
        'Vakıf Standı Yönetici Odası',
        'TSK',
        'Basın',
        'Gönüllü - Bursiyer',
        'Yönetim Ofisi / Kriz Masası',
        'T3 Ofis',
        'Teknofest Robolig Yarışması',
        'Teknofest KKTC Araştırma Yarışması',
        'VIP Salon',
        'TRT Kulis Ana Sahne / Muhabir',
        'Bilim Pavyonu',
        'Pilot Event'
    ]
    
    # Tüm günler
    gunler = ['1. Gün', '2. Gün', '3. Gün', '4. Gün']
    
    # Kontrol zamanları
    kontrol_zamanlari = {
        '09.00': {'limit_saat': '10.00'},
        '14.30': {'limit_saat': '15.30'}
    }
    
    # Sonuç verilerini tutacak yapı
    # İlk key: gün, ikinci key: kontrol zamanı, üçüncü key: alan
    gun_kontrol_verileri = {}
    
    # 14.00 saati (sabah ve öğleden sonra ayırımı için)
    ayirim_saati = datetime.strptime('14.00', '%H.%M').time()
    
    # Her gün ve kontrol zamanı için verileri hazırla
    for gun in gunler:
        gun_kontrol_verileri[gun] = {}
        
        for kontrol_zamani, zaman_bilgisi in kontrol_zamanlari.items():
            gun_kontrol_verileri[gun][kontrol_zamani] = {}
            
            limit_saat = datetime.strptime(zaman_bilgisi['limit_saat'], '%H.%M').time()
            
            # Her alan için
            for alan in alanlar:
                # Varsayılan değerler
                gun_kontrol_verileri[gun][kontrol_zamani][alan] = {
                    'gelme_durumu': None,
                    'gelme_saati': None,
                    'gelme_durumu_renk': 'bg-light text-muted',  # Varsayılan renk - veri yok
                    'gelme_saati_renk': 'bg-light text-muted'  # Varsayılan renk - veri yok
                }
                
                # Veritabanından o gün ve alan için girişleri çek
                veriler = GonulluDurumVeriler.objects.filter(
                    gun=gun,
                    alan=alan
                )
                
                # Kontrol saati için uygun veriyi seç
                if kontrol_zamani == '09.00':
                    # 09.00 kontrolü için 14.00'dan önce girilen tüm kayıtları al
                    filtrelenmis_veriler = veriler.filter(saat__lt=ayirim_saati).order_by('-submitteddate', '-submittedtime')
                else:
                    # 14.30 kontrolü için 14.00 ve sonrası girilen tüm kayıtları al
                    filtrelenmis_veriler = veriler.filter(saat__gte=ayirim_saati).order_by('-submitteddate', '-submittedtime')
                
                # Eğer ilgili saat aralığında veri yoksa, o alan için hiçbir veri gösterme
                if not filtrelenmis_veriler.exists():
                    continue
                
                # Filtrelenmiş verilerin ilki (en son girilen kayıt)
                veri = filtrelenmis_veriler.first()  
                
                # Gelme durumu
                if veri.catering_durum == 'var':
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_durumu'] = 'Geldi'
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_durumu_renk'] = 'bg-success text-white'
                    
                    # Gelme saati
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_saati'] = veri.saat.strftime('%H.%M')
                    
                    # Gelme saati kontrolü (10.00'dan önce/sonra veya 15.30'dan önce/sonra)
                    saat, dakika = veri.saat.hour, veri.saat.minute
                    limit_saat, limit_dakika = map(int, zaman_bilgisi['limit_saat'].split('.'))
                    
                    if (saat < limit_saat) or (saat == limit_saat and dakika < limit_dakika):
                        gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_saati_renk'] = 'bg-success text-white'
                    else:
                        gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_saati_renk'] = 'bg-danger text-white'
                else:
                    # Gelme durumu "Gelmedi" ise
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_durumu'] = 'Gelmedi'
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_durumu_renk'] = 'bg-danger text-white'
                    
                    # Gelme saati - gelmedi ise de saati göster ama kutu kırmızı olsun
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_saati'] = veri.saat.strftime('%H.%M')
                    gun_kontrol_verileri[gun][kontrol_zamani][alan]['gelme_saati_renk'] = 'bg-danger text-white'
    
    # Excel indirme işlemi
    if 'excel' in request.GET:
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        # Stil eklemek için gerekli modülleri içe aktar
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
        
        # Renk tanımlamaları
        baslik_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")  # Gri
        success_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Yeşil
        danger_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Kırmızı
        light_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")    # Açık gri
        
        # Yazı renkleri
        beyaz_font = Font(color="FFFFFF", bold=True)
        siyah_font = Font(color="000000", bold=True)
        
        # Hücre kenarlıkları
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Her gün için ayrı bir sayfa oluştur
        for gun in gunler:
            # Sayfa oluştur ve isimlendir
            ws = wb.create_sheet(title=gun)
            
            # Başlık satırı 1
            header_row1 = [""]  # İlk sütun boş
            current_col = 2     # Sütun B'den başla
            
            for alan in alanlar:
                # Alan adını ekle ve hücre birleştirmesi yap
                ws.cell(row=1, column=current_col, value=alan)
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col+2)
                
                # Birleştirilmiş hücrenin formatını ayarla
                merged_cell = ws.cell(row=1, column=current_col)
                merged_cell.fill = baslik_fill
                merged_cell.font = beyaz_font
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.border = thin_border
                
                # Sonraki alan için sütun indeksini güncelle
                current_col += 3
            
            # İlk sütun başlık stilini uygula
            ws.cell(row=1, column=1).fill = baslik_fill
            ws.cell(row=1, column=1).font = beyaz_font
            ws.cell(row=1, column=1).border = thin_border
            
            # Başlık satırı 2 - Alt başlıklar
            current_col = 2  # Sütun B'den başla
            
            # İlk sütunu boş bırak (A sütunu - kontrol zamanları için)
            ws.cell(row=2, column=1, value="")
            
            # Her alan için alt başlıkları ekle (Gelme Durumu, Geldiği Saat, Resim)
            for _ in alanlar:
                ws.cell(row=2, column=current_col, value="Gelme Durumu")
                ws.cell(row=2, column=current_col+1, value="Geldiği Saat")
                ws.cell(row=2, column=current_col+2, value="Resim")
                current_col += 3
            
            # Başlık satırı 2'ye stil uygula
            for col in range(1, len(alanlar) * 3 + 2):
                cell = ws.cell(row=2, column=col)
                cell.fill = baslik_fill
                cell.font = beyaz_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Veri satırları
            row_idx = 3
            
            # 09.00 kontrolü satırı ekle
            row = ["09.00 kontrolü"]
            col_values = []
            
            for alan in alanlar:
                if gun in gun_kontrol_verileri and '09.00' in gun_kontrol_verileri[gun] and alan in gun_kontrol_verileri[gun]['09.00']:
                    veri = gun_kontrol_verileri[gun]['09.00'][alan]
                    # Veri varsa değerini, yoksa boş string ekle
                    gelme_durumu = veri['gelme_durumu'] if veri['gelme_durumu'] else ""
                    gelme_saati = veri['gelme_saati'] if veri['gelme_saati'] else ""
                    
                    # Renk kodlarını al
                    gelme_durumu_renk = veri['gelme_durumu_renk']
                    gelme_saati_renk = veri['gelme_saati_renk']
                    
                    # Fotoğraf linkini ekle - önce veritabanından ilgili kaydı bul
                    fotograf_linkleri = ""
                    if gelme_durumu:  # Eğer gelme durumu varsa fotoğraf da arayabiliriz
                        try:
                            # İlgili gün, alan ve saate göre veritabanında kayıt ara
                            kayit = GonulluDurumVeriler.objects.filter(
                                gun=gun,
                                alan=alan,
                                saat=datetime.strptime(gelme_saati, '%H.%M').time() if gelme_saati else None
                            ).order_by('-submitteddate', '-submittedtime').first()
                            
                            if kayit:
                                # Önce yeni çoklu fotoğraf sistemindeki fotoğrafları kontrol et
                                fotograflar = list(kayit.fotograflar.all())
                                if fotograflar:
                                    # Tüm fotoğraf linklerini numaralı olarak göster
                                    fotograf_linkleri = ""  # Değişkeni boş string ile başlat
                                    for i, foto in enumerate(fotograflar, 1):
                                        fotograf_linkleri += f"{i}. {foto.get_fotograf_url()}\n"
                                    fotograf_linkleri = fotograf_linkleri.strip()
                        except Exception as e:
                            # Hata durumunda sessizce geç - excelde fotoğraf görünmeyecek
                            pass
                    else:
                        # Eski tek fotoğraf sistemini kontrol et
                        if kayit and kayit.fotograf:
                            fotograf_linkleri = f"1. {kayit.get_fotograf_url()}"
                    col_values.extend([
                        (gelme_durumu, gelme_durumu_renk),
                        (gelme_saati, gelme_saati_renk),
                        (fotograf_linkleri, 'bg-light text-muted')  # Fotoğraf linkleri
                    ])
                else:
                    # Veri yoksa boş hücreler ekle (üç sütun için)
                    col_values.extend([
                        ("", 'bg-light text-muted'),
                        ("", 'bg-light text-muted'),
                        ("", 'bg-light text-muted')
                    ])
            
            # Sadece değerleri ekle
            ws.append(row + [val[0] for val in col_values])
            
            # Şimdi stilleri uygula
            # İlk sütun başlık stilini uygula (Kontrol saati)
            ws.cell(row=row_idx, column=1).fill = baslik_fill
            ws.cell(row=row_idx, column=1).font = beyaz_font
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=1).border = thin_border
            
            # Verilere stil uygula
            for col_idx, (_, renk) in enumerate(col_values, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Renkleri uygula
                if 'bg-success' in renk:
                    cell.fill = success_fill
                    cell.font = beyaz_font
                elif 'bg-danger' in renk:
                    cell.fill = danger_fill
                    cell.font = beyaz_font
                else:
                    cell.fill = light_fill
                    cell.font = siyah_font
                
                # Eğer fotoğraf sütunu ise hücre formatını ayarla
                if (col_idx - 2) % 3 == 2:  # 3 sütunluk yapıda her 3. sütun (resim sütunu)
                    # URL olduğundan hiperlink olarak ayarla
                    value = cell.value
                    if value and value.strip():
                        # Numaralı listeye göre düzenlenmiş URL'leri koruyalım
                        urls = value.split('\n')
                        
                        # İlk URL'yi hücreye hiperlink olarak ekle (eğer varsa)
                        for url in urls:
                            # "1. http://..." formatından URL kısmını al
                            link_match = re.search(r'\d+\.\s*(https?://\S+)', url)
                            if link_match:
                                cell.hyperlink = link_match.group(1)
                                break
                        
                        # URL'leri olduğu gibi göster, kısaltma yapma
                        cell.value = value
                        
                        # URL sütunları için hücre biçimlendirmesi
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.font = Font(color="0000FF")
            
            row_idx += 1

            # 09.00 ve 14.30 arası tüm verileri ekle (en fazla 9 satır)
            for alan in alanlar:
                # İlgili gün ve alan için tüm verileri al
                sabah_verileri = []
                try:
                    # Saat filtrelemesi ile 09.00-14.00 arası kayıtları al ve saat sırasına göre sırala
                    veriler = GonulluDurumVeriler.objects.filter(
                        gun=gun,
                        alan=alan,
                        saat__lt=ayirim_saati  # 14.00'dan önce
                    ).order_by('saat')  # saat sırasına göre (erkenden geçe)
                    
                    # İlk veri (son eklenen, dashboard'da gösterilen) hariç diğer verileri al
                    if veriler.exists():
                        sabah_verileri = list(veriler)[:-1] if len(veriler) > 1 else []
                except Exception as e:
                    # Hata durumunda boş liste kullan
                    sabah_verileri = []
                
                # Alana göre tüm verileri ekle
                for i, alan_idx in enumerate(range(0, len(alanlar) * 3, 3)):
                    if i == alanlar.index(alan) and sabah_verileri:
                        # Bu alan için veri ekle (en fazla 9 veri)
                        for veri_idx, veri in enumerate(sabah_verileri[:9]):
                            # Satır yoksa yeni satır ekle
                            if row_idx + veri_idx >= len(ws._cells) or row_idx + veri_idx not in ws._cells:
                                # Yeni satır ekle
                                ws.append([""] * (len(alanlar) * 3 + 1))
                            
                            # Hücre değerlerini ekle
                            # Gelme durumu
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).value = "Geldi" if veri.catering_durum == 'var' else "Gelmedi"
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).fill = success_fill if veri.catering_durum == 'var' else danger_fill
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).font = beyaz_font
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).border = thin_border
                            
                            # Gelme saati
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).value = veri.saat.strftime('%H.%M')
                            
                            # Saat kontrolü (10.00'dan önce mi sonra mı)
                            limit_saat = datetime.strptime(kontrol_zamanlari['09.00']['limit_saat'], '%H.%M').time()
                            saat_bg = success_fill if veri.saat < limit_saat else danger_fill
                            
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).fill = saat_bg
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).font = beyaz_font
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).border = thin_border
                            
                            # Fotoğraf linkleri
                            fotograflar = list(veri.fotograflar.all())
                            fotograf_linkleri = ""
                            
                            if fotograflar:
                                for f_idx, foto in enumerate(fotograflar, 1):
                                    fotograf_linkleri += f"{f_idx}. {foto.get_fotograf_url()}\n"
                            elif veri.fotograf:  # Eski tek fotoğraf sistemi
                                fotograf_linkleri = f"1. {veri.get_fotograf_url()}"
                                
                            fotograf_linkleri = fotograf_linkleri.strip()
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).value = fotograf_linkleri
                            
                            # Fotoğraf hücresinin formatı
                            if fotograf_linkleri:
                                # İlk URL'yi hiperlink olarak ayarla
                                urls = fotograf_linkleri.split('\n')
                                for url in urls:
                                    link_match = re.search(r'\d+\.\s*(https?://\S+)', url)
                                    if link_match:
                                        ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).hyperlink = link_match.group(1)
                                        break
                                
                                ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).font = Font(color="0000FF")
                                
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).border = thin_border
            
            # En fazla 9 satır ekleyeceğiz
            row_idx += 9
            
            # 14.30 kontrolü satırı ekle
            row = ["14.30 kontrolü"]
            col_values = []
            
            for alan in alanlar:
                if gun in gun_kontrol_verileri and '14.30' in gun_kontrol_verileri[gun] and alan in gun_kontrol_verileri[gun]['14.30']:
                    veri = gun_kontrol_verileri[gun]['14.30'][alan]
                    # Veri varsa değerini, yoksa boş string ekle
                    gelme_durumu = veri['gelme_durumu'] if veri['gelme_durumu'] else ""
                    gelme_saati = veri['gelme_saati'] if veri['gelme_saati'] else ""
                    
                    # Renk kodlarını al
                    gelme_durumu_renk = veri['gelme_durumu_renk']
                    gelme_saati_renk = veri['gelme_saati_renk']
                    
                    # Fotoğraf linkini ekle - önce veritabanından ilgili kaydı bul
                    fotograf_linkleri = ""
                    if gelme_durumu:  # Eğer gelme durumu varsa fotoğraf da arayabiliriz
                        try:
                            # İlgili gün, alan ve saate göre veritabanında kayıt ara
                            kayit = GonulluDurumVeriler.objects.filter(
                                gun=gun,
                                alan=alan,
                                saat=datetime.strptime(gelme_saati, '%H.%M').time() if gelme_saati else None
                            ).order_by('-submitteddate', '-submittedtime').first()
                            
                            if kayit:
                                # Önce yeni çoklu fotoğraf sistemindeki fotoğrafları kontrol et
                                fotograflar = list(kayit.fotograflar.all())
                                if fotograflar:
                                    # Tüm fotoğraf linklerini numaralı olarak göster
                                    fotograf_linkleri = ""  # Değişkeni boş string ile başlat
                                    for i, foto in enumerate(fotograflar, 1):
                                        fotograf_linkleri += f"{i}. {foto.get_fotograf_url()}\n"
                                    fotograf_linkleri = fotograf_linkleri.strip()
                        except Exception as e:
                            # Hata durumunda sessizce geç - excelde fotoğraf görünmeyecek
                            pass
                    else:
                        # Eski tek fotoğraf sistemini kontrol et
                        if kayit and kayit.fotograf:
                            fotograf_linkleri = f"1. {kayit.get_fotograf_url()}"
                    col_values.extend([
                        (gelme_durumu, gelme_durumu_renk),
                        (gelme_saati, gelme_saati_renk),
                        (fotograf_linkleri, 'bg-light text-muted')  # Fotoğraf linkleri
                    ])
                else:
                    # Veri yoksa boş hücreler ekle (üç sütun için)
                    col_values.extend([
                        ("", 'bg-light text-muted'),
                        ("", 'bg-light text-muted'),
                        ("", 'bg-light text-muted')
                    ])
            
            # Sadece değerleri ekle
            ws.append(row + [val[0] for val in col_values])
            
            # Şimdi stilleri uygula
            # İlk sütun başlık stilini uygula (Kontrol saati)
            ws.cell(row=row_idx, column=1).fill = baslik_fill
            ws.cell(row=row_idx, column=1).font = beyaz_font
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=1).border = thin_border
            
            # Verilere stil uygula
            for col_idx, (_, renk) in enumerate(col_values, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Renkleri uygula
                if 'bg-success' in renk:
                    cell.fill = success_fill
                    cell.font = beyaz_font
                elif 'bg-danger' in renk:
                    cell.fill = danger_fill
                    cell.font = beyaz_font
                else:
                    cell.fill = light_fill
                    cell.font = siyah_font
                
                # Eğer fotoğraf sütunu ise hücre formatını ayarla
                if (col_idx - 2) % 3 == 2:  # 3 sütunluk yapıda her 3. sütun (resim sütunu)
                    # URL olduğundan hiperlink olarak ayarla
                    value = cell.value
                    if value and value.strip():
                        # Numaralı listeye göre düzenlenmiş URL'leri koruyalım
                        urls = value.split('\n')
                        
                        # İlk URL'yi hücreye hiperlink olarak ekle (eğer varsa)
                        for url in urls:
                            # "1. http://..." formatından URL kısmını al
                            link_match = re.search(r'\d+\.\s*(https?://\S+)', url)
                            if link_match:
                                cell.hyperlink = link_match.group(1)
                                break
                        
                        # URL'leri olduğu gibi göster, kısaltma yapma
                        cell.value = value
                        
                        # URL sütunları için hücre biçimlendirmesi
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.font = Font(color="0000FF")
            
            row_idx += 1
            
            # 14.30 sonrası tüm verileri ekle
            for alan in alanlar:
                # İlgili gün ve alan için tüm verileri al
                aksam_verileri = []
                try:
                    # Saat filtrelemesi ile 14.00 ve sonrası kayıtları al ve saat sırasına göre sırala
                    veriler = GonulluDurumVeriler.objects.filter(
                        gun=gun,
                        alan=alan,
                        saat__gte=ayirim_saati  # 14.00 ve sonrası
                    ).order_by('saat')  # saat sırasına göre (erkenden geçe)
                    
                    # İlk veri (son eklenen, dashboard'da gösterilen) hariç diğer verileri al
                    if veriler.exists():
                        aksam_verileri = list(veriler)[:-1] if len(veriler) > 1 else []
                except Exception as e:
                    # Hata durumunda boş liste kullan
                    aksam_verileri = []
                
                # Alana göre tüm verileri ekle
                for i, alan_idx in enumerate(range(0, len(alanlar) * 3, 3)):
                    if i == alanlar.index(alan) and aksam_verileri:
                        # Bu alan için veri ekle (tüm verileri)
                        for veri_idx, veri in enumerate(aksam_verileri):
                            # Satır yoksa yeni satır ekle
                            if row_idx + veri_idx >= len(ws._cells) or row_idx + veri_idx not in ws._cells:
                                # Yeni satır ekle
                                ws.append([""] * (len(alanlar) * 3 + 1))
                            
                            # Hücre değerlerini ekle
                            # Gelme durumu
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).value = "Geldi" if veri.catering_durum == 'var' else "Gelmedi"
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).fill = success_fill if veri.catering_durum == 'var' else danger_fill
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).font = beyaz_font
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=row_idx + veri_idx, column=2 + alan_idx).border = thin_border
                            
                            # Gelme saati
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).value = veri.saat.strftime('%H.%M')
                            
                            # Saat kontrolü (15.30'dan önce mi sonra mı)
                            limit_saat = datetime.strptime(kontrol_zamanlari['14.30']['limit_saat'], '%H.%M').time()
                            saat_bg = success_fill if veri.saat < limit_saat else danger_fill
                            
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).fill = saat_bg
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).font = beyaz_font
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).alignment = Alignment(horizontal='center', vertical='center')
                            ws.cell(row=row_idx + veri_idx, column=3 + alan_idx).border = thin_border
                            
                            # Fotoğraf linkleri
                            fotograflar = list(veri.fotograflar.all())
                            fotograf_linkleri = ""
                            
                            if fotograflar:
                                for f_idx, foto in enumerate(fotograflar, 1):
                                    fotograf_linkleri += f"{f_idx}. {foto.get_fotograf_url()}\n"
                            elif veri.fotograf:  # Eski tek fotoğraf sistemi
                                fotograf_linkleri = f"1. {veri.get_fotograf_url()}"
                                
                            fotograf_linkleri = fotograf_linkleri.strip()
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).value = fotograf_linkleri
                            
                            # Fotoğraf hücresinin formatı
                            if fotograf_linkleri:
                                # İlk URL'yi hiperlink olarak ayarla
                                urls = fotograf_linkleri.split('\n')
                                for url in urls:
                                    link_match = re.search(r'\d+\.\s*(https?://\S+)', url)
                                    if link_match:
                                        ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).hyperlink = link_match.group(1)
                                        break
                                
                                ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).font = Font(color="0000FF")
                                
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            ws.cell(row=row_idx + veri_idx, column=4 + alan_idx).border = thin_border
            
        # Sütun genişliklerini ve satır yüksekliklerini ayarla
        # Excel'de sütun genişlikleri karakter cinsinden, satır yükseklikleri nokta (point) cinsindendir
        ws.column_dimensions['A'].width = 24
        
        # Her alan 3 sütundan oluşuyor (Gelme Durumu, Geldiği Saat, Resim)
        for i in range(len(alanlar) * 3):
            col_letter = get_column_letter(i + 2)  # 2'den başla çünkü A sütunu zaten ayarlandı
            
            # Resim sütunlarını daha geniş yap (her 3. sütun)
            if (i % 3) == 2:
                ws.column_dimensions[col_letter].width = 30  # Resim linki için daha geniş
            else:
                ws.column_dimensions[col_letter].width = 20  # Diğer sütunlar için normal genişlik
        
        # Satır yüksekliklerini arttır (fotoğraf linklerini gösterebilmek için)
        for i in range(1, 5):  
            row_height = 60 if i > 2 else 40  # Veri satırları için daha yüksek
            ws.row_dimensions[i].height = row_height
        
        # İlk sayfayı sil (varsayılan oluşturulan)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Excel dosyasını bellekte oluştur
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # HTTP yanıtı olarak döndür
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Gonullu_Durum_Raporu.xlsx"'
        return response
    
    context = {
        'secilen_gun': secilen_gun,
        'gunler': gunler,
        'alanlar': alanlar,
        'kontrol_zamanlari': kontrol_zamanlari,
        'gun_kontrol_verileri': gun_kontrol_verileri
    }
    
    return render(request, 'dashboard/gonullu_durum_raporu.html', context)